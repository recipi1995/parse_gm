from config import *
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from time import sleep
from transliterate import translit
from time import time
import shutil
import requests
import json
import os

def download_image(url,path):
    print("Загрузка фото .....")
    head, tail = os.path.split(path)
    if not os.path.exists(head):
        os.makedirs(head)
    if url.startswith('//'):
        url = url.replace('//','https://').replace('http://','https://')
    img_data = requests.get(url).content
    with open(path, 'wb') as handler:
        handler.write(img_data)

def get_photos(driver,country,city,name,place_id):
    temp = driver.execute_script('return JSON.stringify(JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][171])')
    temp = json.loads(temp)
    k = 0
    blocked = ['Недавние','Меню','Кофе']
    folder_name = name.replace('"','').replace("'",'')
    result = []
    try:  
        for i in range(len(temp[0])):
            if temp[0][i][2] not in blocked and k < 4:
                k+=1
                download_image(temp[0][i][3][0][6][0],f'./data/photos/{country}/{city}/{folder_name} {place_id}/img_{k}.png')
                result.append(f'img_{k}.png')
                print(f"Фото {i} +")
    except:
        print("фото закончились")
        pass
    return result

def get_info(driver):
    print("Получение информации о компании")
    location_data = {}    
    location_data["location_lat"] = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[4][0][1]')
    location_data["location_lng"] = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[4][0][2]')
    location_data["formatted_address"] = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][39]')
    try:
        a = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][4][3][0]')
        location_data["place_id"] = list(map(str,a.split('&')))[0].replace('https://search.google.com/local/reviews?placeid=','')
        print("Получен place_id")
    except:
        print("Не найден place_id")
        pass
    try:
        temp_status = driver.execute_script('return document.querySelector("#pane > div > div > div > div > div > div > div ").innerText')
        if 'Временно закрыто' in temp_status:
            location_data['business_status'] = 'CLOSED_TEMPORARILY'
            print("Бизнес статус CLOSED_TEMPORARILY")
        else:            
            location_data['business_status'] = 'OPERATIONAL'
            print("Бизнес статус OPERATIONAL")
            try:
                a = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][203]')
                a = a[0]
                a.sort(key = lambda x: x[1])
                location_data["opening_hours"] = ', '.join([item[3][0][0] for item in a])
                print("часы работы")
                location_data["weekday_text"] = ', '.join([item[0] for item in a])
                print("дни работы")
            except:
                print('Нет часов работы')
                pass
    except:
        print("Нет бизнес статуса")
        pass
    try:
        a = driver.execute_script('return document.querySelector("#pane > div > div.widget-pane-content > div > div > div > div > div > div > div > div:nth-child(2) > span > span:nth-child(1) > button").innerText')
        location_data["types_name"] = a
        print("Оригинальное имя организации получено")
    except:
        print("У организации не удалось получить оригинальное имя")
        pass
    try:
        a = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][100][1]')
        location_data["types"] = ', '.join([item[0] for item in a])
        print("Тип организации получен")
    except:
        print("У организации не удалось получить тип")
        pass
    try:
        location_data["rating"] = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][4][7]')
        print("Рейтинг организации получен")
    except:
        print("Не удалось получить рейтинг организации")
        pass
    try:
        print("Получен рейтинг по отзывам")
        location_data["user_ratings_total"] = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][4][8]')
    except:
        print("Не удалось получить рейтинг по отзывам")
        pass
    try:
        location_data["website"] = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][7][0]')
        print("Веб сайт получен")
    except:
        print("Не удалось получить веб сайт")
        pass
    try:
        location_data["formatted_phone_number"] = driver.execute_script('return JSON.parse(window.APP_INITIALIZATION_STATE[3][6].slice(5))[6][178][0][3]')
        print("Телефон получен")
    except:
        print("Не удалось получить телефон")
        pass
    print("Данные по организации собраны")
    return location_data

def load_url(driver,url):
    print(f"Загрузка страницы организации: {url}")
    c=0
    searchbox_loading = 'loading show-loading'
    driver.get(url)
    sleep(2)
    while searchbox_loading in driver.find_element(By.ID,'searchbox').get_attribute('class'):
        c=c+1
        sleep(0.8)
        print(f"load_url waiting count {c}")
    sleep(1.5)

def parse_places(driver,typ,city,grabbed_names):
    print("Получение организаций по городу")
    searchbox_loading = 'loading show-loading'
    scrollbox = '"#pane > div > div > div > div > div.section-scrollbox > div.section-scrollbox"'
    result = []
    sleep(1.8)
    container = driver.find_element(By.ID,'pane')
    soup = BeautifulSoup(container.get_attribute('innerHTML'),'html.parser')
    sleep(1.8)
    restaurants = soup.find('div',{'aria-label':f'Результаты по запросу "{typ.replace("+"," ")} {city}"'}).find_all('a')
    nxt = driver.find_element(By.CSS_SELECTOR,'[aria-label="Следующая страница"]')
    all_grabbed = nxt.get_attribute('disabled')
    places = {}
    while not all_grabbed:
        try:
            while searchbox_loading in driver.find_element(By.ID,'searchbox').get_attribute('class'):
                sleep(0.8)
            for i in range(6):
                print(f"Скрол {i}")
                driver.execute_script(f'document.querySelector({scrollbox}).scrollTo(0,document.querySelector({scrollbox}).scrollHeight)')
                sleep(0.8)
            sleep(2.5)
            container = driver.find_element(By.ID,'pane')
            soup = BeautifulSoup(container.get_attribute('innerHTML'),'html.parser')
            restaurants = soup.find('div',{'aria-label':f'Результаты по запросу "{typ.replace("+"," ")} {city}"'}).find_all('a')
            nxt = driver.find_element(By.CSS_SELECTOR,'[aria-label="Следующая страница"]')
            all_grabbed = nxt.get_attribute('disabled')    
            print('Найдено организаций',len(restaurants))
            for item in restaurants:
                temp_name = translit(item['aria-label'].replace(' ','_'),'ru',reversed=True)
                if temp_name not in grabbed_names:
                    result.append({'name':temp_name,'href':item['href'], 'original_name':item['aria-label']})
                    grabbed_names.add(temp_name)                
            if not all_grabbed:
                nxt.click()
                sleep(2)
        except Exception as e:
            print(e)
            all_grabbed = True
    return result

def json_to_excel(city, country):
    print("Запись в exceel")
    try:
        os.mkdir(f"{os.getcwd()}\\data\\excel\\{country}")
    except FileExistsError:
        print("FileEXist")
        pass
    shutil.copyfile(f"{os.getcwd()}\\data\\excel\\template.xlsx", f"{os.getcwd()}\\data\\excel\\{country}\\{city}.xlsx")
    filename = f'{os.getcwd()}\data\excel\{country}\{city}.xlsx'
    wb = load_workbook(filename)
    sheet_ranges = wb['Лист1']
    file = open(f'{city}_updated.json','r', encoding='utf8')
    data = json.load(file)
    file.close()
    c=0
    g=0
    for j in range(len(data)):
        item = data[j]
        for i in range(1,18):        
            if sheet_ranges[f'{get_column_letter(i)}1'].value in item:
                c=c+1
                print(f"jsontoexcel false {c}")
                if item[sheet_ranges[f'{get_column_letter(i)}1'].value] != None:
                    sheet_ranges[f'{get_column_letter(i)}{j+2}'].value = str(item[sheet_ranges[f'{get_column_letter(i)}1'].value])
                else:
                    sheet_ranges[f'{get_column_letter(i)}{j+2}'].value = 'NA'
            else:
                g=g+1
                print(f"jsontoexcel true {g}")
                if sheet_ranges[f'{get_column_letter(i)}1'].value == 'url':
                    sheet_ranges[f'{get_column_letter(i)}{j+2}'].value = str(item['href'])
                else:
                    sheet_ranges[f'{get_column_letter(i)}{j+2}'].value = 'NA'
    wb.save(filename = filename)

def get_all_data(city,country, mode):
    mode = mode
    print(f'Working in {mode} mode')
    # gather | collect | excel_write
    objects = []
    country_value = country
    city_value = city

    if mode != 'excel_write':
        print("open browser")
        options = Options()
        options.binary_location = BINARY_LOCATION
        options.headless = False
        options.add_experimental_option('w3c', True)
        service = Service(BROWSER)
        driver = webdriver.Chrome(service=service,options=options)
        driver.maximize_window()

    if mode == 'gather':
        print("Start in gather mode")
        start = time()
        names = set([])
        for place_value in PLACE_TYPE:
            start_temp = time()
            if len(objects)<40000:
                #load_url(driver,f'https://www.google.com/maps/search/{place_value}+Париж/')
                load_url(driver,f'https://www.google.com/maps/search/{place_value}+{city_value}+,+{country_value}/')
                try:
                    #print(f"Переменная Objects:{objects}")
                    print(f"Переменные передаваемые в parce_places: {place_value}, {city_value+' , '+country_value}, {names}")
                    objects += parse_places(driver,place_value,city_value+' , '+country_value,names)
                except Exception as e:
                    print(e)
                    pass
            print(f'{place_value} is done! time for work is {time()-start_temp}')
        with open(f'{city_value}.json', 'w', encoding='utf8') as fout:
            json.dump(objects, fout, ensure_ascii=False)
        print(f'{mode} is done! time for work is {time()-start}')


    elif mode == 'collect':
        print("collect")
        start = time()
        file = open(f'{city_value}.json','r', encoding='utf8')
        data = json.load(file)
        file.close()
        names = set([])
        for item in data:
            try:
                load_url(driver,item['href'])
                item_data = get_info(driver)
                item_data['name'] = translit(item['name'],'ru',reversed=True).replace(' ','_')
                item_data['url'] = item['href']
                item_data['original_name'] = item['original_name']
                item = item_data
            except:
                pass
            #print('place_id' in item, item['business_status'] == 'OPERATIONAL')
            if 'place_id' in item and item['business_status'] == 'OPERATIONAL' and item_data['name'] not in names:
                names.add(item_data['name'])
                try:
                    photos = get_photos(driver,country_value,city_value,item['name'],item_data["place_id"])
                    if len(photos)>0:
                        folder_name = item['name'].replace('"','').replace("'",'')
                        item_data["photos_field"] = f'./{country_value}/{city_value}/{folder_name} {item_data["place_id"]}/'+','.join(photos)
                except Exception as e:
                    print(e)
                objects.append(item_data)
        with open(f'{city_value}_updated.json', 'w', encoding='utf8') as fout:
            json.dump(objects, fout, ensure_ascii=False)
        print(f'{mode} is done! time for work is {time()-start}')
    
    if mode == 'excel_write':
        start = time()
        json_to_excel(city_value, country_value)
        print(f'{mode} is done! time for work is {time()-start}')

    if mode != 'excel_write':
        print('excel_write driver.quit')
        driver.quit()
    return print(f"Данные по {city_value} собраны")


def start_parce(dict_by_cities):
    dict_cities = dict_by_cities 
    for country in dict_cities:
        for city in dict_cities[country]:
            print(city,country)
            get_all_data(city,country, mode="gather")
            get_all_data(city,country, mode="collect")
            get_all_data(city,country, mode="excel_write")
    return print(f"Все данные собраны по городам из списка: {dict_cities}")


if __name__ == "__main__":
    start_parce(DICT_CITIES)


